"""
hh_gradio.py — парсер вакансий hh.ru с веб-интерфейсом (Gradio)

Установка:
    pip install playwright openpyxl gradio
    playwright install chromium

Запуск:
    python3 hh_gradio.py
    Откроется браузер на http://localhost:7860
"""

import threading
import urllib.parse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import gradio as gr
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# НАСТРОЙКИ
# ─────────────────────────────────────────────

SCROLL_PAUSE_MS = 1200
SCROLL_MAX = 200
STABLE_ROUNDS = 5

CITIES = {
    "Вся Россия":       "113",
    "Москва":           "1",
    "Санкт-Петербург":  "2",
    "Екатеринбург":     "3",
    "Новосибирск":      "4",
    "Казань":           "88",
    "Нижний Новгород":  "66",
    "Самара":           "78",
    "Ростов-на-Дону":   "76",
    "Краснодар":        "53",
    "Удалённая работа": "",
}

SESSION_DIR = str(Path.home() / ".hh_session")
OUTPUT_DIR  = Path.home() / "Downloads"
OUTPUT_DIR.mkdir(exist_ok=True)

# Глобальный флаг для остановки
_stop_flag = threading.Event()


# ─────────────────────────────────────────────
# МОДЕЛЬ ДАННЫХ
# ─────────────────────────────────────────────

@dataclass
class Vacancy:
    title:     str = ""
    company:   str = ""
    salary:    str = ""
    city:      str = ""
    published: str = ""
    watchers:  str = ""
    url:       str = ""


# ─────────────────────────────────────────────
# ПАРСИНГ
# ─────────────────────────────────────────────

def scroll_to_bottom(page, log_fn) -> int:
    cards = page.locator('[data-qa="vacancy-serp__vacancy"]')
    stable = 0
    prev = cards.count()
    log_fn(f"  Прокрутка... карточек: {prev}")

    for i in range(1, SCROLL_MAX + 1):
        if _stop_flag.is_set():
            break
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(SCROLL_PAUSE_MS)
        cur = cards.count()
        if cur > prev:
            log_fn(f"  Скролл {i}: +{cur - prev} (итого {cur})")
            prev = cur
            stable = 0
        else:
            stable += 1
            if stable >= STABLE_ROUNDS:
                break
    return prev


def parse_page(page) -> list[Vacancy]:
    page.wait_for_selector('[data-qa="vacancy-serp__vacancy"]', timeout=30_000)
    cards = page.locator('[data-qa="vacancy-serp__vacancy"]')
    result = []

    for i in range(cards.count()):
        if _stop_flag.is_set():
            break
        card = cards.nth(i)
        v = Vacancy()

        t = card.locator('[data-qa="serp-item__title-text"]').first
        v.title = t.inner_text().strip() if t.count() else ""

        a = card.locator('a[data-qa="serp-item__title"]').first
        href = a.get_attribute("href") if a.count() else ""
        v.url = href.split("?")[0] if href else ""

        c = card.locator('[data-qa="vacancy-serp__vacancy-employer-text"]').first
        if not c.count():
            c = card.locator('[data-qa="vacancy-serp__vacancy-employer"]').first
        v.company = c.inner_text().strip() if c.count() else "—"

        s = card.locator('[data-qa="vacancy-serp__vacancy-compensation"]').first
        v.salary = s.inner_text().strip() if s.count() else "не указана"

        ci = card.locator('[data-qa="vacancy-serp__vacancy-address"]').first
        if not ci.count():
            ci = card.locator('[data-qa="vacancy-serp__vacancy-address-text"]').first
        v.city = ci.inner_text().strip().replace("\n", ", ") if ci.count() else "—"

        d = card.locator('[data-qa="vacancy-serp__vacancy-date"]').first
        v.published = d.inner_text().strip() if d.count() else "—"

        w = card.locator('span:has-text("Сейчас смотрят")').first
        v.watchers = w.inner_text().strip() if w.count() else "—"

        result.append(v)
    return result


def save_to_excel(vacancies: list[Vacancy], query: str, path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["№", "Название вакансии", "Компания", "Зарплата", "Город", "Опубликовано", "Смотрят сейчас", "Ссылка"]
    col_widths = [5, 45, 30, 22, 25, 16, 18, 55]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    for ri, v in enumerate(vacancies, 2):
        bg = "EBF3FB" if ri % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", start_color=bg)
        row_data = [ri - 1, v.title, v.company, v.salary, v.city, v.published, v.watchers, v.url]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            if ci == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="1F4E79", underline="single")
        ws.row_dimensions[ri].height = 18

    total_row = len(vacancies) + 2
    ws.cell(row=total_row, column=1, value="Итого:")
    ws.cell(row=total_row, column=2, value=f"{len(vacancies)} вакансий по запросу «{query}»")
    ws.cell(row=total_row, column=7, value=f"Собрано: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    for ci in range(1, 9):
        cell = ws.cell(row=total_row, column=ci)
        cell.font = Font(bold=True, name="Arial", size=9, color="1F4E79")
        cell.fill = PatternFill("solid", start_color="D6E4F0")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(vacancies) + 1}"
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# ОСНОВНАЯ ФУНКЦИЯ — вызывается из Gradio
# ─────────────────────────────────────────────

def run_parser(query: str, city: str, filename: str, limit: int = 100, login_confirmed: bool = False):
    """Генератор — yield отдаёт строки лога в реальном времени."""

    _stop_flag.clear()

    if not query.strip():
        yield "❌ Введи поисковый запрос.", None
        return

    filename = (filename.strip() or "vacancies") + ".xlsx"
    output_path = str(OUTPUT_DIR / filename)
    area_id = CITIES.get(city, "1")

    log_lines = []
    def log(msg):
        log_lines.append(msg)

    try:
        log("=== Запуск браузера ===")
        yield "\n".join(log_lines), None

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=SESSION_DIR,
                headless=False,
            )
            page = context.new_page()

            log("Открываю hh.ru...")
            yield "\n".join(log_lines), None

            page.goto("https://hh.ru/", wait_until="domcontentloaded")
            page.wait_for_timeout(2000)
            log("✅ Браузер открыт, продолжаю.")

            # Формируем URL
            params = {"text": query.strip()}
            if area_id:
                params["area"] = area_id
            if city == "Удалённая работа":
                params["schedule"] = "remote"
            search_url = "https://hh.ru/search/vacancy?" + urllib.parse.urlencode(params)

            log(f"\n=== Поиск: «{query}» / {city} ===")
            log(f"URL: {search_url}")
            yield "\n".join(log_lines), None

            page.goto(search_url, wait_until="domcontentloaded")

            try:
                page.wait_for_selector('[data-qa="vacancy-serp__vacancy"]', timeout=30_000)
            except Exception:
                log("❌ Вакансии не найдены.")
                yield "\n".join(log_lines), None
                context.close()
                return

            log("✅ Выдача загружена.")
            yield "\n".join(log_lines), None

            all_vacancies = []
            page_num = 0

            while not _stop_flag.is_set():
                log(f"\n--- Страница {page_num + 1} ---")
                yield "\n".join(log_lines), None

                scroll_to_bottom(page, log)
                yield "\n".join(log_lines), None

                vacancies_on_page = parse_page(page)
                all_vacancies.extend(vacancies_on_page)
                log(f"  Собрано на странице: {len(vacancies_on_page)} | Итого: {len(all_vacancies)}")
                yield "\n".join(log_lines), None

                if len(all_vacancies) >= limit:
                    log(f"\n  Достигнут лимит {limit} вакансий — стоп.")
                    break

                # Переходим на следующую страницу через URL
                next_url = search_url + f"&page={page_num + 1}"
                page.goto(next_url, wait_until="domcontentloaded")
                page.wait_for_timeout(1500)

                try:
                    page.wait_for_selector('[data-qa="vacancy-serp__vacancy"]', timeout=10_000)
                except Exception:
                    log(f"\n  Страница {page_num + 2} пустая — конец.")
                    break

                if page.locator('[data-qa="vacancy-serp__vacancy"]').count() == 0:
                    log(f"\n  Страница {page_num + 2} пустая — конец.")
                    break

                page_num += 1

            context.close()

            if _stop_flag.is_set():
                log("\n⛔ Остановлено пользователем.")
                yield "\n".join(log_lines), None
                return

            # Сохраняем
            log(f"\n=== Сохранение: {output_path} ===")
            yield "\n".join(log_lines), None

            save_to_excel(all_vacancies, query, output_path)

            log(f"✅ Готово! Вакансий: {len(all_vacancies)}")
            log(f"   Файл: {output_path}")
            yield "\n".join(log_lines), output_path

    except Exception as e:
        log(f"\n❌ Ошибка: {e}")
        yield "\n".join(log_lines), None


def stop_parser():
    _stop_flag.set()
    return "⛔ Остановка после текущей страницы..."


# ─────────────────────────────────────────────
# ИНТЕРФЕЙС GRADIO
# ─────────────────────────────────────────────

with gr.Blocks(title="HH Parser") as app:

    gr.Markdown("# 🔍 HH Parser\nПарсер вакансий с hh.ru → Excel")

    with gr.Row():
        with gr.Column(scale=2):
            query_input = gr.Textbox(
                label="Поисковый запрос",
                placeholder='"crm-маркетолог" OR "retention manager"',
                lines=2,
            )
            city_input = gr.Dropdown(
                label="Город",
                choices=list(CITIES.keys()),
                value="Санкт-Петербург",
            )
            filename_input = gr.Textbox(
                label="Имя файла (без .xlsx)",
                placeholder="vacancies",
                value="vacancies",
            )
            limit_input = gr.Number(
                label="Максимум вакансий",
                value=100,
                precision=0,
                minimum=1,
            )
            gr.Markdown(f"📁 Файл сохранится в: `{OUTPUT_DIR}`")

        with gr.Column(scale=1):
            gr.Markdown("### Авторизация")
            gr.Markdown(
                "При первом запуске браузер откроется и попросит войти в hh.ru вручную.\n\n"
                "После входа нажми **Запустить** — сессия сохранится и при следующих запусках логин не потребуется."
            )

    with gr.Row():
        start_btn = gr.Button("▶ Запустить", variant="primary", scale=3)
        stop_btn  = gr.Button("⏹ Остановить", variant="stop", scale=1)

    log_output  = gr.Textbox(label="Лог", lines=20, interactive=False)
    file_output = gr.File(label="📥 Скачать файл", visible=True)

    start_btn.click(
        fn=run_parser,
        inputs=[query_input, city_input, filename_input, limit_input, gr.State(False)],
        outputs=[log_output, file_output],
    )

    stop_btn.click(
        fn=stop_parser,
        outputs=log_output,
    )

if __name__ == "__main__":
    app.launch(inbrowser=True, theme=gr.themes.Soft())
